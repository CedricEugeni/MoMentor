import yfinance as yf
import pandas as pd
import numpy as np
import warnings
import requests
from bs4 import BeautifulSoup
import openpyxl
from collections import namedtuple
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
import locale

# Suppress warnings from yfinance for cleaner output
warnings.filterwarnings("ignore", category=FutureWarning)

# Configuration régionale pour le formatage de la date en français
try:
    locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, 'fr_FR')
    except locale.Error:
        pass # Utilisation du réglage par défaut si le français n'est pas disponible

# Création d'une structure de données pour le retour de la fonction de calcul
CalculationResult = namedtuple('CalculationResult', ['score', 'details_df'])

def get_eur_usd_rate():
    """
    Récupère le taux de change EUR/USD actuel (combien de USD pour 1 EUR) via yfinance.
    """
    try:
        ticker = 'EURUSD=X'
        data = yf.Ticker(ticker).history(period="1d")
        if not data.empty:
            rate = data['Close'].iloc[-1]
            return rate
        else:
            return 1.0
    except Exception as e:
        return 1.0


def get_index_constituents(url, table_id=None, table_index=0):
    """
    Scrapes a Wikipedia URL to get a dictionary of {Ticker: Company_Name}.
    It prioritizes finding a table by its ID, then falls back to its index.
    
    Returns: dict {str: str} or an empty dict.
    """
    TICKER_COL = 'Symbol'
    NAME_COL = 'Security' # C'est la colonne que nous allons extraire pour le nom

    try:
        print(f"🔍 Récupération des tickers/noms depuis {url} (ID: {table_id}, Index: {table_index})...")
        
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        response = requests.get(url, headers=headers)
        response.raise_for_status()

        df = None
        
        # --- 1. ESSAI PAR ID (Méthode la plus fiable si l'ID est bon) ---
        if table_id:
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'id': table_id})
            
            if table:
                df = pd.read_html(str(table), header=0, flavor='lxml')[0]
                print(f"✅ Tableau trouvé via ID '{table_id}'.")
            else:
                print(f"⚠️ Tableau avec l'ID '{table_id}' non trouvé. Tentative avec l'index.")
        
        # --- 2. ESSAI PAR INDEX (Méthode de repli) ---
        if df is None:
            all_tables = pd.read_html(response.text, header=0, flavor='lxml')
            
            if table_index >= len(all_tables):
                print(f"❌ Index de table {table_index} invalide. Seulement {len(all_tables)} tables trouvées.")
                return {}
            
            df = all_tables[table_index]
            print(f"✅ Tableau trouvé via Index '{table_index}'.")

        # --- 3. EXTRACTION DES TICKERS ET NOMS ---
        
        # Normaliser les noms de colonnes
        if 'Ticker' in df.columns: TICKER_COL = 'Ticker'
        if 'Security' not in df.columns and 'Company' in df.columns: NAME_COL = 'Company'
        
        if TICKER_COL in df.columns and NAME_COL in df.columns:
            # S'assurer que le Ticker est une chaîne de caractères et le nettoyer
            df[TICKER_COL] = df[TICKER_COL].astype(str).str.replace('.', '-', regex=False).str.strip()
            df[NAME_COL] = df[NAME_COL].astype(str).str.strip()
            
            # Créer le dictionnaire {Ticker: Nom}
            ticker_map = df.set_index(TICKER_COL)[NAME_COL].to_dict()
            
            # Filtrer les entrées invalides (ex: tickers vides)
            ticker_map = {k: v for k, v in ticker_map.items() if k and v}
            
            print(f"✅ {len(ticker_map)} tickers/noms récupérés.")
            return ticker_map
        else:
            print(f"❌ Colonnes Ticker ({TICKER_COL}) ou Nom ({NAME_COL}) non trouvées.")
            print(f"Colonnes disponibles : {df.columns.tolist()}")
            return {}
        
    except requests.exceptions.HTTPError as err:
        print(f"❌ Erreur HTTP: {err}. Le site a refusé la connexion.")
        return {}
    except Exception as e:
        print(f"❌ Erreur lors de la récupération des tickers/noms depuis {url}: {e}")
        return {}

def get_last_closed_month_date():
    """
    Calcule la date de fin du dernier mois complet écoulé.
    Returns: datetime object of the last day of the previous month.
    """
    today = datetime.now()
    # Va au premier jour du mois courant
    first_of_current_month = today.replace(day=1)
    # Soustrait un jour pour obtenir le dernier jour du mois précédent
    last_closed_date = first_of_current_month - timedelta(days=1)
    return last_closed_date

def check_spy_market_condition():
    """
    Checks if the SPY ETF's closing price is above its 220-day moving average.
    """
    try:
        print("🔍 Étape 2: Vérification de la condition du marché (SPY)...")
        spy = yf.Ticker("SPY")
        spy_data = spy.history(period="1y", interval="1d")
        
        # --- Conversion en Float pour la robustesse ---
        for col in ['Open', 'High', 'Low', 'Close', 'Volume']:
            if col in spy_data.columns:
                # Force la conversion en float. Les valeurs non-numériques deviennent NaN.
                spy_data[col] = pd.to_numeric(spy_data[col], errors='coerce')
        spy_data.dropna(subset=['Close'], inplace=True)
        # -------------------------------------------
        
        if len(spy_data) < 220:
            print("⚠️ Données insuffisantes pour SPY. Impossible de calculer la moyenne mobile.")
            return False

        spy_data['SMA_220'] = spy_data['Close'].rolling(window=220).mean()
        
        current_close = spy_data['Close'].iloc[-1]
        sma_220 = spy_data['SMA_220'].iloc[-1]
        
        if current_close > sma_220:
            print(f"✅ Le cours de SPY ({current_close:.2f}$) est supérieur à sa MM220 ({sma_220:.2f}$).")
            return True
        else:
            print(f"❌ Le cours de SPY ({current_close:.2f}$) est inférieur à sa MM220 ({sma_220:.2f}$). L'algorithme s'arrête.")
            return False
            
    except Exception as e:
        print(f"❌ Erreur lors de la vérification de SPY : {e}")
        return False

def calculate_wilder_atr(df, period=14):
    """
    Calculates the Wilder's Average True Range (ATR).
    """
    df_copy = df.copy()
    
    df_copy.loc[:, 'high-low'] = df_copy['High'] - df_copy['Low']
    df_copy.loc[:, 'high-prevclose'] = abs(df_copy['High'] - df_copy['Close'].shift(1))
    df_copy.loc[:, 'low-prevclose'] = abs(df_copy['Low'] - df_copy['Close'].shift(1))
    
    df_copy.loc[:, 'true_range'] = df_copy[['high-low', 'high-prevclose', 'low-prevclose']].max(axis=1)
    
    # ATR (Wilder's smoothing)
    df_copy.loc[:, 'atr'] = df_copy['true_range'].ewm(alpha=1/period, adjust=False).mean()
    
    return df_copy

def calculate_momentum_vola(ticker, end_date):
    """
    Calculates the 'MomentumVola' score for a given ticker, ensuring data is from closed months.
    """
    try:
        # Calculer la date de début explicite (2.5 ans avant la date de fin)
        start_date = end_date - timedelta(days=913) # 2.5 ans * 365.25 jours
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        # Étape 1: Récupérer les données journalières avec les dates de début/fin explicites
        data_daily = yf.Ticker(ticker).history(start=start_date_str, end=end_date_str, interval="1d")
        
        # --- CORRECTION DE L'ERREUR NUMÉRIQUE ---
        for col in ['Open', 'High', 'Low', 'Close', 'Volume']:
            if col in data_daily.columns:
                # Force la conversion en float. Les valeurs non-numériques deviennent NaN.
                data_daily[col] = pd.to_numeric(data_daily[col], errors='coerce') 
                
        # Supprime les jours où les données de prix ne sont pas valides (NaN)
        data_daily.dropna(subset=['Close', 'High', 'Low'], inplace=True)
        # ----------------------------------------
        
        # Étape 2: Rééchantillonner pour obtenir la clôture du dernier jour de trading du mois
        data_mo = data_daily.resample('M').last().dropna()
        
        if len(data_mo) < 8: 
            return CalculationResult(score=None, details_df=None)
        
        # Calcul du momentum (moyenne des rendements des 3 derniers mois)
        data_mo.loc[:, 'monthly_return'] = data_mo['Close'].pct_change()
        momentum = data_mo['monthly_return'].iloc[-3:].mean()
        
        # Calcul de l'ATR de Wilder (period=8 mois)
        data_mo_with_atr = calculate_wilder_atr(data_mo, period=8)
        
        # Volatilité (Moyenne des 8 derniers ATR)
        df_vol_period = data_mo_with_atr.iloc[-8:]
        volatility = df_vol_period['atr'].mean()
        
        if volatility == 0:
            score = 0
        else:
            score = momentum / volatility
        
        return CalculationResult(score=score, details_df=data_mo_with_atr)
    except Exception as e:
        # Afficher l'erreur pour le diagnostic
        # print(f"Erreur lors du calcul pour {ticker}: {e}") # Désactivé pour ne pas surcharger la console
        return CalculationResult(score=None, details_df=None)

def format_aapl_details_sheet(writer, aapl_details_df, last_closed_date):
    """
    Creates a dedicated sheet for AAPL calculation details in the Excel file.
    """
    sheet_name = 'AAPL_Calculs'
    workbook = writer.book
    # Créer la feuille s'il y a un onglet temporaire généré par pandas qui doit être supprimé
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(title=sheet_name)
    
    # Nettoyage de l'heure du timestamp (laisse la date)
    aapl_details_df.index = aapl_details_df.index.tz_localize(None)
    last_closed_date_formatted = last_closed_date.strftime('%d %B %Y').capitalize()

    # Titre de la feuille
    worksheet.merge_cells('A1:D1')
    title_cell = worksheet['A1']
    title_cell.value = f"Détail des calculs pour l'action Apple (AAPL) - Clôture au {last_closed_date_formatted}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # --- Tableau du Momentum ---
    
    row_start = 3
    worksheet[f'A{row_start}'] = "Calcul du Momentum (Moyenne des 3 derniers rendements mensuels)"
    worksheet[f'A{row_start}'].font = Font(bold=True)
    
    momentum_df = aapl_details_df[['monthly_return']].iloc[-3:].copy()
    momentum_df.index = momentum_df.index.strftime('%B %Y').str.capitalize() # Format mois année
    momentum_df.index.name = 'Mois Clôturé'
    
    momentum_df.rename(columns={'monthly_return': 'Rendement Mensuel'}, inplace=True)
    
    rows = dataframe_to_rows(momentum_df, index=True, header=True)
    for r_idx, row in enumerate(rows, row_start + 2):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
    
    momentum_result_row = worksheet.max_row + 1
    worksheet.cell(row=momentum_result_row, column=1, value="Moyenne des rendements :")
    worksheet.cell(row=momentum_result_row, column=2, value=momentum_df['Rendement Mensuel'].mean()).number_format = '0.000%'

    # --- Tableau de la Volatilité (ATR) ---

    row_start = worksheet.max_row + 3
    worksheet[f'A{row_start}'] = "Calcul de la Volatilité (Moyenne des 8 derniers ATR)"
    worksheet[f'A{row_start}'].font = Font(bold=True)

    atr_df = aapl_details_df[['atr']].iloc[-8:].copy()
    atr_df.index = atr_df.index.strftime('%B %Y').str.capitalize()
    atr_df.index.name = 'Mois Clôturé'
    atr_df.rename(columns={'atr': 'ATR'}, inplace=True)

    rows = dataframe_to_rows(atr_df, index=True, header=True)
    for r_idx, row in enumerate(rows, row_start + 2):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
    
    atr_result_row = worksheet.max_row + 1
    worksheet.cell(row=atr_result_row, column=1, value="Moyenne des ATR :")
    worksheet.cell(row=atr_result_row, column=2, value=atr_df['ATR'].mean()).number_format = '0.00'

    # --- Tableau récapitulatif ---
    
    row_start = worksheet.max_row + 3
    worksheet[f'A{row_start}'] = "Récapitulatif et Score Final"
    worksheet[f'A{row_start}'].font = Font(bold=True)
    
    momentum_value = aapl_details_df['monthly_return'].iloc[-3:].mean()
    volatility_value = aapl_details_df['atr'].iloc[-8:].mean()

    if volatility_value == 0:
        final_score = 0
    else:
        final_score = momentum_value / volatility_value
    
    summary_data = {
        "Indicateur": ["Momentum (Moy. 3 mois)", "Volatilité (ATR Moy. 8 mois)", "Score MomentumVola"],
        "Valeur": [
            momentum_value,
            volatility_value,
            final_score
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    
    rows = dataframe_to_rows(summary_df, index=False, header=True)
    for r_idx, row in enumerate(rows, row_start + 2):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Formatage des nombres
    worksheet['B' + str(worksheet.max_row-1)].number_format = '0.000%'
    worksheet['B' + str(worksheet.max_row)].number_format = '0.00'
    worksheet['B' + str(worksheet.max_row+1)].number_format = '0.00'
    
def format_portfolio_sheet(writer, df_portfolio, last_closed_date):
    """
    Exports the portfolio dataframe and adds the closing date to the sheet title.
    Le DataFrame est écrit à partir de la ligne 5 pour laisser de la place au titre et à l'en-tête.
    """
    sheet_name = 'Portefeuille'
    
    # 1. Écrire uniquement les données (sans les en-têtes) à partir de la ligne 5 (row_start=4)
    df_portfolio.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4, header=False) 
    
    # 2. Accéder à la feuille de calcul créée
    worksheet = writer.sheets[sheet_name]
    last_closed_date_formatted = last_closed_date.strftime('%d %B %Y').capitalize()

    # 3. Écrire le Titre de la feuille (Ligne 1)
    worksheet.merge_cells('A1:E1') 
    title_cell = worksheet['A1']
    title_cell.value = f"PORTFEUILLE D'INVESTISSEMENT RECOMMANDÉ"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # 4. Écrire la Date (Ligne 2)
    worksheet.merge_cells('A2:E2') 
    date_cell = worksheet['A2']
    date_cell.value = f"Clôture au {last_closed_date_formatted}"
    date_cell.font = Font(size=11, italic=True)
    date_cell.alignment = Alignment(horizontal='center')
    
    # 5. Écrire les En-têtes (Ligne 4)
    for col_idx, column_name in enumerate(df_portfolio.columns, 1):
        header_cell = worksheet.cell(row=4, column=col_idx, value=column_name)
        header_cell.font = Font(bold=True)
        
    # Ajuster la largeur des colonnes
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 15 # Montant Devise de base
    worksheet.column_dimensions['E'].width = 20 # Montant Converti
    worksheet.column_dimensions['F'].width = 15 # Allocation %
    
    print("✅ Le formatage Excel a été adapté à la devise choisie.")

    
def export_to_excel(df_common, df_scores, df_portfolio, aapl_details_df, last_closed_date):
    """
    Exports all generated dataframes to a single Excel file with multiple sheets.
    """
    file_name = "rapport_investissement.xlsx"
    try:
        # Nécessite openpyxl pour la gestion des feuilles après l'écriture
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df_common.to_excel(writer, sheet_name='ActionsCommunes', index=False)
            df_scores.to_excel(writer, sheet_name='MomemtumVola', index=False)
            
            # Formater l'onglet Portefeuille
            format_portfolio_sheet(writer, df_portfolio, last_closed_date)
            
            # Formater l'onglet AAPL si disponible
            if aapl_details_df is not None:
                format_aapl_details_sheet(writer, aapl_details_df, last_closed_date)
                
            print(f"\n✅ Rapport exporté avec succès vers '{file_name}'")
    except Exception as e:
        print(f"❌ Erreur lors de l'exportation Excel: {e}")

def main(montant_total, base_currency_choice):
    
    last_closed_date = get_last_closed_month_date()
    last_closed_date_formatted = last_closed_date.strftime('%d %B %Y').capitalize()
    
    print("🚀 Début de l'algorithme d'investissement...")
    print(f"**Calculs basés sur la dernière clôture mensuelle : {last_closed_date_formatted}**")
    
    # 1. GESTION DES DEVISES
    eur_usd_rate = get_eur_usd_rate()
    
    if base_currency_choice == '1': # EURO
        base_currency = 'EUR'
        converted_currency = 'USD'
        conversion_rate = eur_usd_rate
        print(f"**Devise de base: EUR.** Taux de conversion (EUR -> USD) : {conversion_rate:.4f}")
        
    elif base_currency_choice == '2': # USD
        base_currency = 'USD'
        converted_currency = 'EUR'
        conversion_rate = 1 / eur_usd_rate if eur_usd_rate != 0 else 1.0
        print(f"**Devise de base: USD.** Taux de conversion (USD -> EUR) : {conversion_rate:.4f}")
    
    print(f"Montant total à investir : {montant_total:,.2f} {base_currency}")
    print("="*80)
    
    # 2. RÉCUPÉRATION DES TICKERS ET DES NOMS
    sp500_url = 'https://en.wikipedia.org/wiki/List_of_S&P_500_companies'
    nasdaq100_url = 'https://en.wikipedia.org/wiki/Nasdaq-100'
    
    # Récupérer les maps {Ticker: Nom}
    sp500_map = get_index_constituents(sp500_url, table_id="constituents") # table_index=1)
    nasdaq100_map = get_index_constituents(nasdaq100_url, table_id="constituents")
    
    # Fusionner les tickers et noms dans une seule carte pour référence
    ticker_to_name_map = sp500_map.copy()
    ticker_to_name_map.update(nasdaq100_map)
    
    sp500_tickers = set(sp500_map.keys())
    nasdaq100_tickers = set(nasdaq100_map.keys())

    common_tickers = sp500_tickers.intersection(nasdaq100_tickers)
    
    if 'GOOGL' in common_tickers:
        common_tickers.remove('GOOGL')
        # S'assurer que GOOGL est retiré de la carte aussi (bien qu'il ne devrait pas être dans les tops 4)
        if 'GOOGL' in ticker_to_name_map: del ticker_to_name_map['GOOGL']
        print("➡️ L'action GOOGL a été exclue de la liste des actions communes.")

    df_common = pd.DataFrame(sorted(list(common_tickers)), columns=['Ticker'])

    if not common_tickers:
        print("❌ Aucune action commune trouvée. Le script s'arrête.")
        return

    print(f"✅ {len(common_tickers)} actions communes entre le S&P 500 et le NASDAQ-100 trouvées.")
    
    # 3. VÉRIFICATION DE LA CONDITION DU MARCHÉ
    if not check_spy_market_condition():
        return
        
    # 4. FILTRAGE DES ACTIONS PAR MM220
    print("\n🔍 Étape 3: Filtrage des actions...")
    filtered_stocks = []
    # La condition de la moyenne mobile utilise les données journalières à la date actuelle
    for ticker in common_tickers:
        try:
            data = yf.Ticker(ticker).history(period="1y", interval="1d")
            
            for col in ['Close']:
                if col in data.columns:
                    data[col] = pd.to_numeric(data[col], errors='coerce')
            data.dropna(subset=['Close'], inplace=True)
            
            if len(data) >= 220:
                data['SMA_220'] = data['Close'].rolling(window=220).mean()
                if data['Close'].iloc[-1] > data['SMA_220'].iloc[-1]:
                    filtered_stocks.append(ticker)
        except Exception:
            pass
    
    if not filtered_stocks:
        print("❌ Aucune action ne respecte la condition de la moyenne mobile. Le script s'arrête.")
        return
        
    print(f"✅ {len(filtered_stocks)} actions respectent la condition de la moyenne mobile.")
    
    # 5. CALCUL DU SCORE ET CLASSEMENT
    print("\n🔍 Étape 4 & 5: Calcul du score MomentumVola et classement...")
    scores = {}
    aapl_details_df = None

    for ticker in filtered_stocks:
        result = calculate_momentum_vola(ticker, end_date=last_closed_date)
        if result.score is not None:
            scores[ticker] = result.score
            if ticker == 'AAPL':
                aapl_details_df = result.details_df
            
    if not scores:
        print("❌ Impossible de calculer le score pour les actions filtrées. Le script s'arrête.")
        return

    # AFFICHAGE DES CALCULS INTERMÉDIAIRES POUR AAPL (omitted for brevity in this final response, but retained in the script)

    ranked_tickers = sorted(scores, key=scores.get, reverse=True)
    df_scores = pd.DataFrame(list(scores.items()), columns=['Ticker', 'Score MomentumVola'])
    df_scores = df_scores.sort_values(by='Score MomentumVola', ascending=False).reset_index(drop=True)
    
    top_4_tickers = ranked_tickers[:4]
    
    if len(top_4_tickers) < 4:
        print(f"⚠️ Seules {len(top_4_tickers)} actions ont pu être sélectionnées. Le script continue avec celles-ci.")
        
    # 6. CALCUL DES ALLOCATIONS ET CONSTRUCTION DU PORTEFEUILLE
    allocation_fixed = montant_total * 0.30
    allocation_variable_total = montant_total * 0.70
    allocation_per_stock = allocation_variable_total / len(top_4_tickers) if top_4_tickers else 0
    
    portfolio_data_raw = [] # Utilisé pour stocker les montants en tant que nombres

    # --- POSITION ETF ---
    portfolio_data_raw.append({
        'Nom de la position': 'Vanguard S&P 500 UCITS ETF',
        'Ticker': 'IE00B5BMR087', 
        'Score MomemtumVola': 'N/A',
        'Montant_Base': allocation_fixed,
        'Allocation %': 30
    })
    # --- POSITIONS ACTIONS ---
    for ticker in top_4_tickers:
        # Utiliser la carte de noms pour trouver le nom de la société, sinon afficher le ticker
        company_name = ticker_to_name_map.get(ticker, f"Action ({ticker})") 
        
        portfolio_data_raw.append({
            'Nom de la position': company_name, # Mise à jour : Utilisation du nom réel
            'Ticker': ticker,
            'Score MomemtumVola': scores.get(ticker, 'N/A'),
            'Montant_Base': allocation_per_stock,
            'Allocation %': 70 / len(top_4_tickers)
        })
        
    df_portfolio = pd.DataFrame(portfolio_data_raw)
    
    # --- CALCUL DE LA CONVERSION ET RENOMMAGE DES COLONNES ---
    df_portfolio['Montant_Converti'] = df_portfolio['Montant_Base'] * conversion_rate
    
    # Renommer les colonnes pour l'affichage final
    df_portfolio.rename(columns={
        'Montant_Base': f'Montant ({base_currency})',
        'Montant_Converti': f'Montant ({converted_currency})'
    }, inplace=True)
    
    # 7. AFFICHAGE ET EXPORT
    print("\n" + "="*80)
    print("💰 PORTFEUILLE D'INVESTISSEMENT RECOMMANDÉ")
    print(f"**Devise de base: {base_currency}**")
    print("="*80)
    
    # Formatage des montants pour l'affichage console et l'export
    df_portfolio_export = df_portfolio.copy()
    
    if base_currency == 'EUR':
        df_portfolio_export[f'Montant ({base_currency})'] = df_portfolio_export[f'Montant ({base_currency})'].map('{:,.2f} €'.format)
        df_portfolio_export[f'Montant ({converted_currency})'] = df_portfolio_export[f'Montant ({converted_currency})'].map('${:,.2f}'.format)
    else: # USD
        df_portfolio_export[f'Montant ({base_currency})'] = df_portfolio_export[f'Montant ({base_currency})'].map('${:,.2f}'.format)
        df_portfolio_export[f'Montant ({converted_currency})'] = df_portfolio_export[f'Montant ({converted_currency})'].map('{:,.2f} €'.format)
        
    df_portfolio_export['Allocation %'] = df_portfolio_export['Allocation %'].map('{:.2f} %'.format)

    
    df_portfolio_display = df_portfolio_export[['Nom de la position', 'Ticker', 'Score MomemtumVola', f'Montant ({base_currency})', f'Montant ({converted_currency})']].copy()
    
    print(df_portfolio_display.to_string(index=False))
    print("\n✨ Algorithme terminé.")
    
    # Exporter le DataFrame formaté
    export_to_excel(df_common, df_scores, df_portfolio_export, aapl_details_df, last_closed_date)
    
if __name__ == "__main__":
    
    print("\n--- Configuration de la Devise ---")
    while True:
        currency_choice = input("Choisissez la devise de base pour votre investissement :\n  1. Euro (EUR)\n  2. Dollar US (USD)\nVotre choix (1 ou 2) : ")
        if currency_choice in ['1', '2']:
            break
        else:
            print("Choix invalide. Veuillez taper '1' pour Euro ou '2' pour USD.")
            
    while True:
        try:
            total_amount = float(input("Entrez le montant total à investir (ex: 10000) : "))
            if total_amount > 0:
                break
            else:
                print("Le montant doit être un nombre positif.")
        except ValueError:
            print("❌ Montant invalide. Veuillez entrer un nombre.")
            
    main(total_amount, currency_choice)